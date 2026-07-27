import os
import sys
import shutil
import csv
from datetime import datetime
from flask import Flask, render_template, request, jsonify
import openpyxl
from openpyxl.styles import PatternFill, Font

app = Flask(__name__)

COLOR_OK_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
COLOR_ERR_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

DEFAULT_WORKBOOK_PATH = r"C:\Users\hana\Documents\GitHub\webtools\daily-checker\蔵書点検チェックツール4.xlsm"

def ensure_trailing_slash(path):
    if path and not path.endswith("\\") and not path.endswith("/"):
        return path + "\\"
    return path

def read_config():
    """Excelのメインシートから設定値を読み込む"""
    wb = openpyxl.load_workbook(DEFAULT_WORKBOOK_PATH, data_only=True)
    ws = wb.worksheets[0]
    
    b3_val = ws["B3"].value
    b3_date_str = ""
    if isinstance(b3_val, datetime):
        b3_date_str = b3_val.strftime("%Y-%m-%d")
    elif b3_val:
        b3_date_str = str(b3_val).split(" ")[0].replace("/", "-")
        
    config = {
        "path1": ws["A1"].value or "",
        "file1": ws["B1"].value or "",
        "path2": ws["A2"].value or "",
        "file2": ws["B2"].value or "",
        "backupPath": ws["A3"].value or "",
        "dateStr": b3_date_str,
        "workFileName": ws["B5"].value or "",
        "textFilePath": ws["A9"].value or "",
        "destPath": ws["A21"].value or "",
        "historyPath": ws["A25"].value or "",
        "historyFile": ws["B25"].value or "",
        # 件数は初期表示では常に 0 (実行時に動的更新)
        "count_E1": ws["E1"].value if ws["E1"].value is not None else 0,
        "count_E2": ws["E2"].value if ws["E2"].value is not None else 0,
        "count_F1": ws["F1"].value if ws["F1"].value is not None else 0,
        "count_F2": ws["F2"].value if ws["F2"].value is not None else 0,
        "count_G1": ws["G1"].value if ws["G1"].value is not None else 0,
        "count_G9": ws["G9"].value if ws["G9"].value is not None else 0
    }
    wb.close()
    return config

def write_config_cell(cell_name, value):
    wb = openpyxl.load_workbook(DEFAULT_WORKBOOK_PATH)
    ws = wb.worksheets[0]
    ws[cell_name] = value
    wb.save(DEFAULT_WORKBOOK_PATH)
    wb.close()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/config', methods=['GET'])
def get_config_api():
    try:
        cfg = read_config()
        return jsonify({"success": True, "config": cfg})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/step1', methods=['POST'])
def step1_backup():
    try:
        cfg = read_config()
        b3_str = request.json.get("dateStr") or cfg["dateStr"]
        if not b3_str:
            return jsonify({"success": False, "error": "B3セルに有効な日付がありません。"})
        
        b3_dt = datetime.strptime(b3_str, "%Y-%m-%d")
        date_str = b3_dt.strftime("%Y%m%d")
        
        path1 = ensure_trailing_slash(cfg["path1"])
        file1 = cfg["file1"].replace(".xlsx", "") + date_str + ".xlsx"
        path2 = ensure_trailing_slash(cfg["path2"])
        file2 = cfg["file2"].replace(".xlsx", "") + date_str + ".xlsx"
        backup_path = ensure_trailing_slash(cfg["backupPath"])
        
        full1 = os.path.join(path1, file1)
        full2 = os.path.join(path2, file2)
        backup1 = os.path.join(backup_path, file1)
        backup2 = os.path.join(backup_path, file2)
        
        if not os.path.exists(backup_path):
            os.makedirs(backup_path, exist_ok=True)
            
        exists1 = os.path.exists(full1)
        exists2 = os.path.exists(full2)
        
        if not exists1 and not exists2:
            return jsonify({"success": False, "error": f"対象ファイルが見つかりません:\n1: {full1}\n2: {full2}"})
            
        if exists1:
            shutil.copy2(full1, backup1)
        if exists2:
            shutil.copy2(full2, backup2)
            
        work_filename = f"書庫点検済みチェックリスト（点検済）_{date_str}.xlsx"
        work_full = os.path.join(backup_path, work_filename)
        
        if os.path.exists(work_full):
            os.remove(work_full)
            
        if exists1:
            shutil.copy2(backup1, work_full)
        elif exists2:
            shutil.copy2(backup2, work_full)
            
        # クリア & 書込 (件数セルを完全にクリア/0化)
        wb = openpyxl.load_workbook(DEFAULT_WORKBOOK_PATH)
        ws = wb.worksheets[0]
        ws["B5"] = work_filename
        for cell in ["E1", "E2", "F1", "F2", "G1", "G9"]:
            ws[cell] = 0
        wb.save(DEFAULT_WORKBOOK_PATH)
        wb.close()
        
        return jsonify({
            "success": True, 
            "message": f"バックアップ完了。作業ファイル作成: {work_filename}",
            "workFileName": work_filename
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/step2', methods=['POST'])
def step2_merge():
    try:
        cfg = read_config()
        b3_str = cfg["dateStr"]
        if not b3_str:
            return jsonify({"success": False, "error": "B3セルに有効な日付がありません。"})
        b3_dt = datetime.strptime(b3_str, "%Y-%m-%d")
        date_str = b3_dt.strftime("%Y%m%d")
        
        backup_path = ensure_trailing_slash(cfg["backupPath"])
        check_filename = cfg["workFileName"]
        check_full = os.path.join(backup_path, check_filename)
        
        if not check_filename or not os.path.exists(check_full):
            return jsonify({"success": False, "error": "作業用ファイルが見つかりません。先に処理1を実行してください。"})
            
        backup1 = os.path.join(backup_path, cfg["file1"].replace(".xlsx", "") + date_str + ".xlsx")
        backup2 = os.path.join(backup_path, cfg["file2"].replace(".xlsx", "") + date_str + ".xlsx")
        exists1 = os.path.exists(backup1)
        exists2 = os.path.exists(backup2)
        
        check_wb = openpyxl.load_workbook(check_full)
        check_ws = check_wb.worksheets[0]
        
        last_check_row = check_ws.max_row
        count_check_date = 0
        
        for r in range(1, last_check_row + 1):
            val_a = check_ws.cell(row=r, column=1).value
            dt_a = None
            if isinstance(val_a, datetime):
                dt_a = val_a.date()
            elif isinstance(val_a, str):
                try:
                    dt_a = datetime.strptime(val_a.split(" ")[0].replace("/", "-"), "%Y-%m-%d").date()
                except:
                    pass
            if dt_a == b3_dt.date():
                count_check_date += 1
                
        count_id_date = 0
        last_id_row = 0
        
        if exists1 and exists2:
            id_wb = openpyxl.load_workbook(backup2, data_only=True)
            id_ws = id_wb.worksheets[0]
            last_id_row = id_ws.max_row
            
            for r in range(1, last_id_row + 1):
                val_a = id_ws.cell(row=r, column=1).value
                dt_a = None
                if isinstance(val_a, datetime):
                    dt_a = val_a.date()
                elif isinstance(val_a, str):
                    try:
                        dt_a = datetime.strptime(val_a.split(" ")[0].replace("/", "-"), "%Y-%m-%d").date()
                    except:
                        pass
                if dt_a == b3_dt.date():
                    count_id_date += 1
                    new_r = check_ws.max_row + 1
                    check_ws.cell(row=new_r, column=1).value = val_a
                    check_ws.cell(row=new_r, column=2).value = id_ws.cell(row=r, column=2).value
                    check_ws.cell(row=new_r, column=3).value = id_ws.cell(row=r, column=3).value
            id_wb.close()
            
        header = [check_ws.cell(row=1, column=c).value for c in range(1, check_ws.max_column + 1)]
        
        data_rows = []
        for r in range(2, check_ws.max_row + 1):
            val_a = check_ws.cell(row=r, column=1).value
            dt_a = None
            if isinstance(val_a, datetime):
                dt_a = val_a.date()
            elif isinstance(val_a, str):
                try:
                    dt_a = datetime.strptime(val_a.split(" ")[0].replace("/", "-"), "%Y-%m-%d").date()
                except:
                    pass
            if dt_a == b3_dt.date():
                row_vals = [check_ws.cell(row=r, column=c).value for c in range(1, check_ws.max_column + 1)]
                data_rows.append(row_vals)
                
        data_rows.sort(key=lambda x: str(x[1]) if x[1] is not None else "")
        
        check_ws.delete_rows(1, check_ws.max_row)
        check_ws.append(header)
        for dr in data_rows:
            check_ws.append(dr)
            
        check_row_count = len(data_rows)
        check_wb.save(check_full)
        check_wb.close()
        
        wb = openpyxl.load_workbook(DEFAULT_WORKBOOK_PATH)
        ws = wb.worksheets[0]
        ws["E1"] = last_check_row
        ws["F1"] = count_check_date
        ws["E2"] = last_id_row
        ws["F2"] = count_id_date
        ws["G1"] = check_row_count
        wb.save(DEFAULT_WORKBOOK_PATH)
        wb.close()
        
        return jsonify({
            "success": True,
            "message": f"リスト統合完了。有効件数: {check_row_count}件",
            "stats": {"E1": last_check_row, "F1": count_check_date, "E2": last_id_row, "F2": count_id_date, "G1": check_row_count}
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/step3', methods=['POST'])
def step3_import():
    try:
        cfg = read_config()
        text_path = cfg["textFilePath"]
        if not text_path or not os.path.exists(text_path):
            return jsonify({"success": False, "error": f"テキストファイルが見つかりません: {text_path}"})
            
        backup_path = ensure_trailing_slash(cfg["backupPath"])
        check_filename = cfg["workFileName"]
        check_full = os.path.join(backup_path, check_filename)
        
        check_wb = openpyxl.load_workbook(check_full)
        check_ws = check_wb.worksheets[0]
        
        imported_count = 0
        with open(text_path, 'r', encoding='utf-8', errors='ignore') as f:
            sample = f.read(2048)
            f.seek(0)
            delimiter = '\t' if '\t' in sample else ','
            reader = csv.reader(f, delimiter=delimiter)
            
            row_idx = 1
            for fields in reader:
                if not fields or not any(fields):
                    continue
                for c_idx, val in enumerate(fields):
                    cell_col = 5 + c_idx
                    check_ws.cell(row=row_idx, column=cell_col).value = val
                row_idx += 1
            imported_count = row_idx - 2 if row_idx >= 2 else 0
            
        check_wb.save(check_full)
        check_wb.close()
        
        write_config_cell("G9", imported_count)
        
        return jsonify({
            "success": True,
            "message": f"テキスト取り込み完了。取り込み件数: {imported_count}件",
            "G9": imported_count
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/step4', methods=['POST'])
def step4_validate():
    try:
        cfg = read_config()
        backup_path = ensure_trailing_slash(cfg["backupPath"])
        check_filename = cfg["workFileName"]
        check_full = os.path.join(backup_path, check_filename)
        
        if not os.path.exists(check_full):
            return jsonify({"success": False, "error": "チェック用ファイルが見つかりません。"})
            
        check_wb = openpyxl.load_workbook(check_full)
        check_ws = check_wb.worksheets[0]
        
        check_ws["D1"] = "チェック"
        check_ws["A1"].fill = COLOR_OK_FILL
        check_ws["B1"].fill = COLOR_OK_FILL
        
        last_row = check_ws.max_row
        
        has_err = {col: False for col in ["B", "E", "F", "G", "H", "I", "K", "L", "M", "N", "P", "Q"]}
        
        valid_H = ["図・開架・参考図書", "図・書庫", "図・開架・図書", "図・開架・ＰＢ"]
        valid_I = ["書庫内図書", "参考図書", "学生図書"]
        
        for i in range(2, last_row + 1):
            val_b = check_ws.cell(row=i, column=2).value
            val_e = check_ws.cell(row=i, column=5).value
            
            b_num = float(val_b) if val_b is not None and str(val_b).strip().isdigit() else None
            e_num = float(val_e) if val_e is not None and str(val_e).strip().isdigit() else None
            
            diff_val = "0"
            if b_num is not None and e_num is not None:
                diff_val = str(int(b_num - e_num))
            check_ws.cell(row=i, column=4).value = f"=B{i}-E{i}"
            
            if not str(val_b or "").strip() or diff_val != "0":
                check_ws.cell(row=i, column=2).fill = COLOR_ERR_FILL
                has_err["B"] = True
                
            if not str(val_e or "").strip():
                check_ws.cell(row=i, column=5).fill = COLOR_ERR_FILL
                has_err["E"] = True
                
            val_f = str(check_ws.cell(row=i, column=6).value or "").strip()
            if not val_f:
                check_ws.cell(row=i, column=6).fill = COLOR_ERR_FILL
                has_err["F"] = True
                
            val_a = check_ws.cell(row=i, column=1).value
            date_a = ""
            if isinstance(val_a, datetime):
                date_a = val_a.strftime("%Y%m%d")
            elif val_a:
                date_a = str(val_a).replace("-", "").replace("/", "").split(" ")[0]
                
            val_g = str(check_ws.cell(row=i, column=7).value or "").strip().replace("/", "").replace("-", "")
            if not val_g or date_a != val_g:
                check_ws.cell(row=i, column=7).fill = COLOR_ERR_FILL
                has_err["G"] = True
                
            val_h = str(check_ws.cell(row=i, column=8).value or "").strip()
            if val_h not in valid_H:
                check_ws.cell(row=i, column=8).fill = COLOR_ERR_FILL
                has_err["H"] = True
                
            val_i = str(check_ws.cell(row=i, column=9).value or "").strip()
            if val_i not in valid_I:
                check_ws.cell(row=i, column=9).fill = COLOR_ERR_FILL
                has_err["I"] = True
                
            val_k = str(check_ws.cell(row=i, column=11).value or "").strip()
            if val_k and val_k != "通常":
                check_ws.cell(row=i, column=11).fill = COLOR_ERR_FILL
                has_err["K"] = True
                
            val_l = str(check_ws.cell(row=i, column=12).value or "").strip()
            if val_l:
                check_ws.cell(row=i, column=12).fill = COLOR_ERR_FILL
                has_err["L"] = True
                
            val_m = str(check_ws.cell(row=i, column=13).value or "").strip()
            if val_m and val_m != "通常":
                check_ws.cell(row=i, column=13).fill = COLOR_ERR_FILL
                has_err["M"] = True
                
            val_n = str(check_ws.cell(row=i, column=14).value or "").strip()
            if val_n and val_n != "通常":
                check_ws.cell(row=i, column=14).fill = COLOR_ERR_FILL
                has_err["N"] = True
                
            val_p = str(check_ws.cell(row=i, column=16).value or "").strip()
            if not val_p:
                check_ws.cell(row=i, column=16).fill = COLOR_ERR_FILL
                has_err["P"] = True
                
            val_q = str(check_ws.cell(row=i, column=17).value or "").strip()
            if not val_q:
                check_ws.cell(row=i, column=17).fill = COLOR_ERR_FILL
                has_err["Q"] = True

        col_map = {"B": 2, "E": 5, "F": 6, "G": 7, "H": 8, "I": 9, "K": 11, "L": 12, "M": 13, "N": 14, "P": 16, "Q": 17}
        for col_letter, c_num in col_map.items():
            check_ws.cell(row=1, column=c_num).fill = COLOR_ERR_FILL if has_err[col_letter] else COLOR_OK_FILL
            
        for c_num in [10, 18, 19]:
            check_ws.cell(row=1, column=c_num).fill = COLOR_OK_FILL
            
        check_wb.save(check_full)
        check_wb.close()
        
        err_cols = [k for k, v in has_err.items() if v]
        
        return jsonify({
            "success": True,
            "message": f"検証完了。{'エラーあり列: ' + ', '.join(err_cols) if err_cols else '全検証項目正常 (エラーなし)'}",
            "errorColumns": err_cols
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/step5', methods=['POST'])
def step5_copy():
    try:
        cfg = read_config()
        src_path = ensure_trailing_slash(cfg["backupPath"])
        dest_path = ensure_trailing_slash(cfg["destPath"])
        file_name = cfg["workFileName"]
        
        src_full = os.path.join(src_path, file_name)
        dest_full = os.path.join(dest_path, file_name)
        
        if not os.path.exists(src_full):
            return jsonify({"success": False, "error": f"元ファイルが見つかりません: {src_full}"})
            
        if not os.path.exists(dest_path):
            os.makedirs(dest_path, exist_ok=True)
            
        if os.path.exists(dest_full):
            return jsonify({"success": False, "error": f"コピー先に同名ファイルが既に存在します: {dest_full}"})
            
        shutil.copy2(src_full, dest_full)
        
        try:
            os.startfile(dest_path)
        except Exception:
            pass
            
        return jsonify({
            "success": True,
            "message": f"ファイルをコピーし、保存先フォルダを開きました:\n{dest_full}"
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/step6', methods=['POST'])
def step6_history():
    try:
        cfg = read_config()
        b3_str = cfg["dateStr"]
        if not b3_str:
            return jsonify({"success": False, "error": "B3セルに有効な日付がありません。"})
        b3_dt = datetime.strptime(b3_str, "%Y-%m-%d").date()
        
        history_dir = ensure_trailing_slash(cfg["historyPath"])
        history_file = cfg["historyFile"]
        history_full = os.path.join(history_dir, history_file)
        
        if not os.path.exists(history_full):
            return jsonify({"success": False, "error": f"履歴ファイルが開けません: {history_full}"})
            
        wb_hist = openpyxl.load_workbook(history_full)
        ws_hist = wb_hist.worksheets[0]
        
        found_row = 0
        for r in range(1, ws_hist.max_row + 1):
            val_a = ws_hist.cell(row=r, column=1).value
            dt_a = None
            if isinstance(val_a, datetime):
                dt_a = val_a.date()
            elif isinstance(val_a, str):
                try:
                    dt_a = datetime.strptime(val_a.split(" ")[0].replace("/", "-"), "%Y-%m-%d").date()
                except:
                    pass
            if dt_a == b3_dt:
                found_row = r
                break
                
        if found_row == 0:
            wb_hist.close()
            return jsonify({"success": False, "error": f"履歴ファイル内で指定日付 ({b3_str}) の行が見つかりません。"})
            
        ws_hist.cell(row=found_row, column=7).value = cfg["count_E2"]
        ws_hist.cell(row=found_row, column=10).value = cfg["count_E1"]
        ws_hist.cell(row=found_row, column=13).value = cfg["count_G9"]
        
        warnings = []
        if str(ws_hist.cell(row=found_row, column=8).value) == str(cfg["count_F2"]):
            ws_hist.cell(row=found_row, column=8).font = Font(bold=True)
        else:
            warnings.append("H列とF2の値が異なります")
            
        if str(ws_hist.cell(row=found_row, column=11).value) == str(cfg["count_F1"]):
            ws_hist.cell(row=found_row, column=11).font = Font(bold=True)
        else:
            warnings.append("K列とF1の値が異なります")
            
        if str(ws_hist.cell(row=found_row, column=12).value) == str(cfg["count_G9"]):
            ws_hist.cell(row=found_row, column=12).font = Font(bold=True)
        else:
            warnings.append("L列とG9の値が異なります")
            
        wb_hist.save(history_full)
        wb_hist.close()
        
        msg = f"履歴ファイルへの記録が正常に完了しました。"
        if warnings:
            msg += f"\n注記: {', '.join(warnings)}"
            
        return jsonify({"success": True, "message": msg, "warnings": warnings})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

if __name__ == '__main__':
    print("Starting Daily Checker Local App on http://127.0.0.1:5000")
    app.run(host='127.0.0.1', port=5000, debug=False)
